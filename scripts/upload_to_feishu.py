import os
import zipfile
import xml.etree.ElementTree as ET
import requests
from openpyxl import load_workbook
import shutil
import re

# Configuration
EXCEL_FILE = '/home/lzwc/project/warehouse/scripts/资产主表，存储资产的基本信息和状态_线上_数据导出 (1).xlsx'
TOKEN = 't-g1042beyCDRUK67GQ2UG7NNWX7B7HIEABQ7YKKTM'
API_URL = 'https://open.feishu.cn/open-apis/im/v1/images'
TEMP_DIR = '/home/lzwc/project/warehouse/scripts/temp_xlsx_extract'

def parse_cell_images_map(temp_dir):
    """
    Parses xl/cellimages.xml and xl/_rels/cellimages.xml.rels to map image IDs (from DISPIMG) to filenames.
    Returns: dict { 'ID_...': 'image1.png' }
    """
    cellimages_path = os.path.join(temp_dir, 'xl', 'cellimages.xml')
    rels_path = os.path.join(temp_dir, 'xl', '_rels', 'cellimages.xml.rels')
    
    if not os.path.exists(cellimages_path) or not os.path.exists(rels_path):
        print("No cellimages.xml found.")
        return {}

    # 1. Parse Relationships (rId -> Target)
    # Target is like "media/image1.png"
    ns_rels = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
    tree = ET.parse(rels_path)
    root = tree.getroot()
    
    rid_to_target = {}
    for rel in root.findall('r:Relationship', ns_rels):
        rid = rel.get('Id')
        target = rel.get('Target')
        rid_to_target[rid] = target

    # 2. Parse Cell Images (ID -> rId)
    # Namespace for WPS custom data might be tricky, usually defined in root
    # The xml snippet showed: xmlns:etc="http://www.wps.cn/officeDocument/2017/etCustomData"
    # and <xdr:cNvPr id="3" name="ID_..."/> and <a:blip r:embed="rId1"/>
    
    # We can use regex or namespace-aware parsing. Regex is often more robust for these specific IDs if namespaces vary.
    # But let's try XML parsing.
    
    with open(cellimages_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Simple regex extraction to avoid namespace hell
    # Looking for name="ID_..." ... r:embed="rId..."
    # The structure is <etc:cellImage> ... <xdr:cNvPr ... name="ID_XX" ...> ... <a:blip r:embed="rIdX"/>
    
    img_map = {}
    
    # Split by cellImage to process each image block
    blocks = content.split('</etc:cellImage>')
    
    for block in blocks:
        id_match = re.search(r'name="(ID_[^"]+)"', block)
        rid_match = re.search(r'r:embed="(rId[^"]+)"', block)
        
        if id_match and rid_match:
            img_id = id_match.group(1)
            rid = rid_match.group(1)
            target = rid_to_target.get(rid)
            if target:
                # Target is usually "media/image1.png". We need just the filename or full path relative to xl
                # The unzip structure puts 'xl/media/image1.png'
                # if target starts with 'media/', we use it.
                img_map[img_id] = target
                
    return img_map

def upload_image(file_path):
    """
    Uploads image to Feishu and returns img_key.
    """
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return None

    headers = {
        'Authorization': f'Bearer {TOKEN}'
    }
    data = {
        'image_type': 'message'
    }
    try:
        with open(file_path, 'rb') as f:
            files = {'image': f}
            response = requests.post(API_URL, headers=headers, data=data, files=files)
            res_json = response.json()
            
            if res_json.get('code') == 0:
                img_key = res_json['data']['image_key']
                print(f"Uploaded {os.path.basename(file_path)} -> {img_key}")
                return img_key
            else:
                print(f"Upload failed for {file_path}: {res_json}")
                return None
    except Exception as e:
        print(f"Exception uploading {file_path}: {e}")
        return None

def main():
    # 1. Extract Excel
    if os.path.exists(TEMP_DIR):
        shutil.rmtree(TEMP_DIR)
    os.makedirs(TEMP_DIR)
    
    print(f"Extracting {EXCEL_FILE}...")
    with zipfile.ZipFile(EXCEL_FILE, 'r') as zip_ref:
        zip_ref.extractall(TEMP_DIR)
        
    # 2. Build Map
    print("Parsing image mapping...")
    id_to_file = parse_cell_images_map(TEMP_DIR)
    print(f"Found {len(id_to_file)} images to upload.")
    
    # 3. Load Workbook
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # 4. Iterate and Replace
    # Assuming column 17 (Q) is the one with images
    IMG_COL_IDX = 17
    
    # Cache uploaded keys to avoid re-uploading same image if used multiple times (unlikely but possible)
    uploaded_keys = {} # { filename: img_key }
    
    updated_count = 0
    
    for row in ws.iter_rows(min_row=2): # Skip header
        cell = row[IMG_COL_IDX - 1] # 0-based index
        val = str(cell.value) if cell.value else ""
        
        if "DISPIMG" in val:
            # Extract ID
            match = re.search(r'DISPIMG\("(ID_[^"]+)"', val)
            if match:
                img_id = match.group(1)
                rel_path = id_to_file.get(img_id)
                
                if rel_path:
                    # Construct full path
                    # rel_path is like "media/image1.png". We need "xl/media/image1.png"
                    full_img_path = os.path.join(TEMP_DIR, 'xl', rel_path)
                    
                    if rel_path in uploaded_keys:
                        img_key = uploaded_keys[rel_path]
                    else:
                        img_key = upload_image(full_img_path)
                        if img_key:
                            uploaded_keys[rel_path] = img_key
                    
                    if img_key:
                        cell.value = img_key
                        updated_count += 1
                else:
                    print(f"Warning: Image ID {img_id} not found in map.")
    
    # 5. Save
    output_file = EXCEL_FILE.replace('.xlsx', '_updated.xlsx')
    wb.save(output_file)
    print(f"Done. Updated {updated_count} cells. Saved to {output_file}")
    
    # Cleanup
    shutil.rmtree(TEMP_DIR)

if __name__ == "__main__":
    main()
